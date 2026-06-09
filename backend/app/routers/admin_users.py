"""Admin user-list export — CSV + XLSX (v3.4.0).

Endpoints
---------
GET /api/v1/admin/users.csv    Streamed CSV (full_name, email, mobile, registered_at)
GET /api/v1/admin/users.xlsx   Same, as an Excel workbook

Access — caller must hold `owner` role in at least one organization. This is
intentionally narrow until we add a global "platform admin" flag; today the
only humans who legitimately need a CSV of all users are the platform's
founding owners.
"""

import csv
import io
import logging
from datetime import datetime

from fastapi import APIRouter, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select

from app.database import DB
from app.models.organization import Membership
from app.models.user import User
from app.security import CurrentUser

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/admin", tags=["Admin"])

_HEADERS = ["Full Name", "Email", "Mobile Number", "Registered At (UTC)"]

async def _require_platform_owner(db, current_user) -> None:
    is_owner = await db.scalar(
        select(Membership.id).where(
            Membership.user_id == current_user.id, Membership.role == "owner"
        )
    )
    if is_owner is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Owner role required for user export.",
        )

async def _fetch_users(db) -> list[User]:
    result = await db.execute(select(User).order_by(User.created_at))
    return list(result.scalars().all())

def _row(u: User) -> list[str]:
    return [
        u.name or "",
        u.email,
        u.phone_number or "",
        u.created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(u.created_at, datetime) else "",
    ]

@router.get("/users.csv")
async def export_users_csv(db: DB, current_user: CurrentUser) -> StreamingResponse:
    await _require_platform_owner(db, current_user)
    users = await _fetch_users(db)

    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(_HEADERS)
    for u in users:
        writer.writerow(_row(u))
    buf.seek(0)
    body = buf.getvalue().encode("utf-8")

    logger.info("admin.users.csv exported count=%d by=%s", len(users), current_user.id)
    return StreamingResponse(
        iter([body]),
        media_type="text/csv",
        headers={
            "Content-Disposition": 'attachment; filename="users.csv"',
            "Content-Length": str(len(body)),
        },
    )

@router.get("/users.xlsx")
async def export_users_xlsx(db: DB, current_user: CurrentUser) -> StreamingResponse:
    await _require_platform_owner(db, current_user)
    users = await _fetch_users(db)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for col, label in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill

    for r, u in enumerate(users, start=2):
        for c, val in enumerate(_row(u), start=1):
            ws.cell(row=r, column=c, value=val)

    for col_letter, width in zip("ABCD", (28, 36, 22, 22), strict=True):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    body = buf.getvalue()

    logger.info("admin.users.xlsx exported count=%d by=%s", len(users), current_user.id)
    return StreamingResponse(
        iter([body]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="users.xlsx"',
            "Content-Length": str(len(body)),
        },
    )
